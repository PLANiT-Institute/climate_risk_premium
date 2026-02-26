#!/usr/bin/env python3
"""
Generate CRP Fact-Check Workbook (Excel).

Reads model results from results/ and data/ directories and produces
a professional, formatted Excel workbook for fact-checking.

Output: results/CRP_Fact_Check_Workbook.xlsx
"""

import csv
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULTS_DIR = os.path.join(BASE_DIR, "results")
DATA_DIR = os.path.join(BASE_DIR, "data", "raw")
SRC_DIR = os.path.join(BASE_DIR, "src", "climada")

OUTPUT_PATH = os.path.join(RESULTS_DIR, "CRP_Fact_Check_Workbook.xlsx")

SCENARIO_CSV = os.path.join(RESULTS_DIR, "scenario_comparison.csv")
CREDIT_CSV = os.path.join(RESULTS_DIR, "credit_ratings.csv")
PLANT_CSV = os.path.join(DATA_DIR, "plant_parameters.csv")

CASHFLOW_FILES = {
    "baseline": os.path.join(RESULTS_DIR, "cashflow_baseline.csv"),
    "moderate_transition": os.path.join(RESULTS_DIR, "cashflow_moderate_transition.csv"),
    "aggressive_transition": os.path.join(RESULTS_DIR, "cashflow_aggressive_transition.csv"),
    "moderate_physical": os.path.join(RESULTS_DIR, "cashflow_moderate_physical.csv"),
    "high_physical": os.path.join(RESULTS_DIR, "cashflow_high_physical.csv"),
    "combined_moderate": os.path.join(RESULTS_DIR, "cashflow_combined_moderate.csv"),
    "combined_aggressive": os.path.join(RESULTS_DIR, "cashflow_combined_aggressive.csv"),
    "low_demand": os.path.join(RESULTS_DIR, "cashflow_low_demand.csv"),
    "severe_drought": os.path.join(RESULTS_DIR, "cashflow_severe_drought.csv"),
    "enhanced_11th_plan": os.path.join(RESULTS_DIR, "cashflow_enhanced_11th_plan.csv"),
    "enhanced_combined": os.path.join(RESULTS_DIR, "cashflow_enhanced_combined.csv"),
}

KEY_SCENARIOS = ["baseline", "aggressive_transition", "enhanced_11th_plan", "low_demand"]
CF_DISPLAY_COLS = ["year", "revenue", "total_costs", "ebitda", "interest_expense", "free_cash_flow", "capacity_factor"]

INVESTMENT_GRADE = {"AAA", "AA", "A", "BBB"}
DISTRESSED_RATINGS = {"CCC", "CC", "C", "D"}

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
RED_FONT = Font(color="C00000")
BOLD_RED_FONT = Font(bold=True, color="C00000")
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def read_csv(path):
    """Read a CSV file and return list of dicts."""
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)


def safe_float(val, default=None):
    """Convert to float, return default on failure."""
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=None):
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def apply_header_style(ws, ncols):
    """Style the first row as header."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_row_styles(ws, start_row, end_row, ncols, neg_cols=None, scenario_col=None, rating_col=None):
    """Apply alternating rows, red for negatives, bold for distressed."""
    neg_cols = neg_cols or []
    for row_idx in range(start_row, end_row + 1):
        fill = ALT_FILL if (row_idx - start_row) % 2 == 1 else WHITE_FILL
        is_distressed = False
        if rating_col is not None:
            rating_val = ws.cell(row=row_idx, column=rating_col).value
            if rating_val in DISTRESSED_RATINGS:
                is_distressed = True
        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Red font for negative numeric values in specified columns
            if col_idx in neg_cols:
                val = cell.value
                if isinstance(val, (int, float)) and val < 0:
                    if is_distressed:
                        cell.font = BOLD_RED_FONT
                    else:
                        cell.font = RED_FONT
                elif is_distressed and scenario_col and col_idx == scenario_col:
                    cell.font = BOLD_FONT
            elif is_distressed and scenario_col and col_idx == scenario_col:
                cell.font = BOLD_FONT


def auto_col_widths(ws, min_width=10, max_width=30):
    """Approximate auto-fit column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def freeze_header(ws):
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def build_summary_sheet(wb, scenario_data):
    ws = wb.create_sheet("Summary")
    headers = [
        "Scenario", "NPV ($M)", "IRR (%)", "Avg DSCR", "Min DSCR",
        "LLCR", "Payback (years)", "Rating", "CRP (bps)",
        "Investment Grade", "Distressed"
    ]
    ws.append(headers)
    apply_header_style(ws, len(headers))

    for row in scenario_data:
        rating = row.get("overall_rating", "")
        ig = "Yes" if rating in INVESTMENT_GRADE else "No"
        dist = "Yes" if rating in DISTRESSED_RATINGS else "No"
        npv = safe_float(row.get("npv_million"))
        irr = safe_float(row.get("irr_pct"))
        avg_dscr = safe_float(row.get("avg_dscr"))
        min_dscr = safe_float(row.get("min_dscr"))
        llcr = safe_float(row.get("llcr"))
        payback = safe_int(row.get("payback_years"))
        crp = safe_float(row.get("crp_bps"))

        ws.append([
            row.get("scenario", ""),
            round(npv, 1) if npv is not None else "",
            round(irr, 2) if irr is not None else "",
            round(avg_dscr, 2) if avg_dscr is not None else "",
            round(min_dscr, 2) if min_dscr is not None else "",
            round(llcr, 2) if llcr is not None else "",
            payback if payback is not None else "",
            rating,
            round(crp, 1) if crp is not None else "",
            ig,
            dist,
        ])

    nrows = ws.max_row
    ncols = len(headers)
    # NPV col=2, IRR col=3, CRP col=9
    neg_cols = [2, 3, 9]
    rating_col = 8  # column H
    scenario_col = 1
    apply_row_styles(ws, 2, nrows, ncols, neg_cols=neg_cols,
                     scenario_col=scenario_col, rating_col=rating_col)

    # Number formats
    for r in range(2, nrows + 1):
        ws.cell(row=r, column=2).number_format = '#,##0.0'
        ws.cell(row=r, column=3).number_format = '0.00'
        ws.cell(row=r, column=4).number_format = '0.00'
        ws.cell(row=r, column=5).number_format = '0.00'
        ws.cell(row=r, column=6).number_format = '0.00'
        ws.cell(row=r, column=9).number_format = '0.0'

    auto_col_widths(ws)
    freeze_header(ws)
    return ws


def build_credit_ratings_sheet(wb, credit_data):
    ws = wb.create_sheet("Credit Ratings")
    if not credit_data:
        ws.append(["No data"])
        return ws
    headers = list(credit_data[0].keys())
    ws.append(headers)
    apply_header_style(ws, len(headers))

    for row in credit_data:
        vals = []
        for h in headers:
            v = row.get(h, "")
            fv = safe_float(v)
            if fv is not None and h != "scenario":
                vals.append(fv)
            else:
                vals.append(v)
        ws.append(vals)

    nrows = ws.max_row
    ncols = len(headers)
    apply_row_styles(ws, 2, nrows, ncols)

    # Format DSCR columns to 2dp
    dscr_col = None
    for i, h in enumerate(headers, 1):
        if h == "dscr":
            dscr_col = i
            break
    if dscr_col:
        for r in range(2, nrows + 1):
            ws.cell(row=r, column=dscr_col).number_format = '0.00'

    auto_col_widths(ws)
    freeze_header(ws)
    return ws


def build_crp_analysis_sheet(wb, scenario_data):
    ws = wb.create_sheet("CRP Analysis")
    headers = [
        "Scenario", "Counterfactual Rating", "Counterfactual Spread (bps)",
        "Scenario Rating", "Scenario Spread (bps)", "Rating Migration",
        "Notch Change", "CRP (bps)", "Debt Spread (bps)",
        "Expected Loss (%)", "NPV Loss ($M)"
    ]
    field_map = [
        "scenario", "counterfactual_rating", "counterfactual_spread_bps",
        "scenario_rating_new", "scenario_spread_bps_new", "rating_migration",
        "notch_change", "crp_bps", "debt_spread_bps",
        "expected_loss_pct", "npv_loss_million"
    ]
    ws.append(headers)
    apply_header_style(ws, len(headers))

    for row in scenario_data:
        vals = []
        for f in field_map:
            v = row.get(f, "")
            fv = safe_float(v)
            if fv is not None and f != "scenario" and f not in (
                "counterfactual_rating", "scenario_rating_new", "rating_migration"
            ):
                vals.append(fv)
            else:
                vals.append(v)
        ws.append(vals)

    nrows = ws.max_row
    ncols = len(headers)
    neg_cols = [7, 8, 11]  # notch change, CRP, NPV loss might be negative
    apply_row_styles(ws, 2, nrows, ncols, neg_cols=neg_cols)

    # Number formats
    for r in range(2, nrows + 1):
        ws.cell(row=r, column=3).number_format = '0.0'
        ws.cell(row=r, column=5).number_format = '0.0'
        ws.cell(row=r, column=8).number_format = '0.0'
        ws.cell(row=r, column=9).number_format = '0'
        ws.cell(row=r, column=10).number_format = '0.00'
        ws.cell(row=r, column=11).number_format = '#,##0.0'

    auto_col_widths(ws)
    freeze_header(ws)
    return ws


def build_cashflows_key_sheet(wb, cashflow_data_dict):
    ws = wb.create_sheet("Cashflows - Key Scenarios")

    # Build side-by-side: Year | baseline cols | aggressive cols | enhanced cols | low_demand cols
    display_labels = {
        "revenue": "Revenue",
        "total_costs": "Total Costs",
        "ebitda": "EBITDA",
        "interest_expense": "Debt Service",
        "free_cash_flow": "Free Cash Flow",
        "capacity_factor": "Capacity Factor",
    }
    value_cols = [c for c in CF_DISPLAY_COLS if c != "year"]

    # Header row 1: scenario group names
    header1 = [""]
    for sc in KEY_SCENARIOS:
        header1.append(sc.replace("_", " ").title())
        header1 += [""] * (len(value_cols) - 1)
    ws.append(header1)

    # Style merged header
    col = 2
    for sc in KEY_SCENARIOS:
        for j in range(len(value_cols)):
            cell = ws.cell(row=1, column=col + j)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1, end_column=col + len(value_cols) - 1
        )
        col += len(value_cols)

    # Header row 2: column sub-headers
    header2 = ["Year"]
    for _ in KEY_SCENARIOS:
        for vc in value_cols:
            header2.append(display_labels.get(vc, vc))
    ws.append(header2)
    ncols = len(header2)
    # Style header row 2
    for c in range(1, ncols + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).border = THIN_BORDER

    # Determine years from baseline
    baseline_rows = cashflow_data_dict.get("baseline", [])
    years = [safe_int(r.get("year")) for r in baseline_rows if safe_int(r.get("year")) is not None]

    # Build lookup: scenario -> year -> row
    lookups = {}
    for sc in KEY_SCENARIOS:
        rows = cashflow_data_dict.get(sc, [])
        lk = {}
        for r in rows:
            y = safe_int(r.get("year"))
            if y is not None:
                lk[y] = r
        lookups[sc] = lk

    data_start_row = 3
    for yr in years:
        row_vals = [yr]
        for sc in KEY_SCENARIOS:
            sc_row = lookups[sc].get(yr, {})
            for vc in value_cols:
                val = safe_float(sc_row.get(vc))
                if val is not None:
                    if vc == "capacity_factor":
                        row_vals.append(round(val, 4))
                    else:
                        row_vals.append(round(val / 1e6, 2))  # Convert to millions
                else:
                    row_vals.append("")
        ws.append(row_vals)

    nrows = ws.max_row
    # Identify columns that can be negative (all value columns except capacity factor)
    neg_cols_set = set()
    col_idx = 2
    for _ in KEY_SCENARIOS:
        for vc in value_cols:
            if vc != "capacity_factor":
                neg_cols_set.add(col_idx)
            col_idx += 1
    apply_row_styles(ws, data_start_row, nrows, ncols, neg_cols=list(neg_cols_set))

    # Number formats
    col_idx = 2
    for _ in KEY_SCENARIOS:
        for vc in value_cols:
            for r in range(data_start_row, nrows + 1):
                cell = ws.cell(row=r, column=col_idx)
                if vc == "capacity_factor":
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0.00'
            col_idx += 1

    auto_col_widths(ws, min_width=12, max_width=18)
    ws.freeze_panes = "A3"
    return ws


def build_all_cashflows_index(wb, cashflow_data_dict):
    ws = wb.create_sheet("All Cashflows Index")
    headers = [
        "Scenario", "Total Revenue ($M)", "Total EBITDA ($M)", "Total FCF ($M)",
        "Avg Annual CF ($M)", "Min Annual CF ($M)", "Max Annual CF ($M)",
        "First Negative CF Year", "Years with Negative CF"
    ]
    ws.append(headers)
    apply_header_style(ws, len(headers))

    scenario_order = [
        "baseline", "moderate_transition", "aggressive_transition",
        "moderate_physical", "high_physical",
        "combined_moderate", "combined_aggressive",
        "low_demand", "severe_drought",
        "enhanced_11th_plan", "enhanced_combined",
    ]

    for sc in scenario_order:
        rows = cashflow_data_dict.get(sc, [])
        if not rows:
            ws.append([sc, "", "", "", "", "", "", "", ""])
            continue

        revenues = [safe_float(r.get("revenue"), 0) for r in rows]
        ebitdas = [safe_float(r.get("ebitda"), 0) for r in rows]
        fcfs = [safe_float(r.get("free_cash_flow"), 0) for r in rows]
        years = [safe_int(r.get("year")) for r in rows]

        total_rev = sum(revenues) / 1e6
        total_ebitda = sum(ebitdas) / 1e6
        total_fcf = sum(fcfs) / 1e6
        avg_cf = (sum(fcfs) / len(fcfs)) / 1e6 if fcfs else 0
        min_cf = min(fcfs) / 1e6 if fcfs else 0
        max_cf = max(fcfs) / 1e6 if fcfs else 0

        neg_years = [y for y, f in zip(years, fcfs) if f < 0 and y is not None]
        first_neg = min(neg_years) if neg_years else "None"
        count_neg = len(neg_years)

        ws.append([
            sc,
            round(total_rev, 1),
            round(total_ebitda, 1),
            round(total_fcf, 1),
            round(avg_cf, 1),
            round(min_cf, 1),
            round(max_cf, 1),
            first_neg,
            count_neg,
        ])

    nrows = ws.max_row
    ncols = len(headers)
    neg_cols = [2, 3, 4, 5, 6, 7]
    apply_row_styles(ws, 2, nrows, ncols, neg_cols=neg_cols)

    for r in range(2, nrows + 1):
        for c in [2, 3, 4, 5, 6, 7]:
            ws.cell(row=r, column=c).number_format = '#,##0.0'

    auto_col_widths(ws)
    freeze_header(ws)
    return ws


def build_assumptions_sheet(wb, plant_data):
    ws = wb.create_sheet("Key Assumptions")

    # Section 1: Plant Parameters
    ws.append(["PLANT PARAMETERS"])
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, size=13, color="1F4E79")

    ws.append(["Parameter", "Value", "Unit", "Description"])
    apply_header_style(ws, 4)
    # Fix: header is row 2 for this section
    for c in range(1, 5):
        cell = ws.cell(row=2, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row in plant_data:
        ws.append([
            row.get("param_name", ""),
            safe_float(row.get("value"), row.get("value", "")),
            row.get("unit", ""),
            row.get("description", ""),
        ])

    plant_end = ws.max_row

    # Blank row
    ws.append([])

    # Section 2: Physical Risk Parameters
    phys_row = ws.max_row + 1
    ws.append(["PHYSICAL RISK PARAMETERS"])
    cell = ws.cell(row=phys_row, column=1)
    cell.font = Font(bold=True, size=13, color="1F4E79")

    ws.append(["Parameter", "Value", "Source"])
    hdr_row = ws.max_row
    for c in range(1, 4):
        cell = ws.cell(row=hdr_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    phys_params = [
        ("Wildfire Base Rate", "0.055%", "Kim et al. (2025) DOI:10.1007/s11069-025-07169-4"),
        ("Flood Base Rate", "0.003%", "Kang & Lee (2024) DOI:10.3390/w16202987"),
        ("SLR Derate per Meter", "0.22%", "Van Vliet et al. (2016) DOI:10.1038/nclimate2903"),
        ("", "", ""),
        ("Wildfire Multiplier RCP4.5 2050", "1.5x", "WWA (2025) interpolated"),
        ("Wildfire Multiplier RCP8.5 2050", "2.0x", "WWA (2025) current attribution"),
        ("Wildfire Multiplier RCP8.5 2100", "4.0x", "WWA (2025) future projection"),
        ("Flood Multiplier RCP8.5 2030", "1.29x", "KSCCR (2024) early century"),
        ("Flood Multiplier RCP8.5 2050", "1.46x", "KSCCR (2024) mid century"),
        ("Flood Multiplier RCP8.5 2100", "2.64x", "npj Climate (2025) DOI:10.1038/s41612-025-01067-z"),
        ("", "", ""),
        ("SLR RCP4.5 2100", "0.25 m", "CMIP6 DOI:10.3390/atmos12010090"),
        ("SLR RCP8.5 2100", "0.63 m", "CMIP6 DOI:10.3390/atmos12010090"),
        ("Compound Multiplier Range", "1.0 - 1.25x", "Zscheischler (2018) framework"),
    ]
    for p in phys_params:
        ws.append(list(p))

    nrows = ws.max_row
    # Apply basic formatting to data rows
    for r in range(3, plant_end + 1):
        fill = ALT_FILL if (r - 3) % 2 == 1 else WHITE_FILL
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER

    phys_data_start = hdr_row + 1
    for r in range(phys_data_start, nrows + 1):
        fill = ALT_FILL if (r - phys_data_start) % 2 == 1 else WHITE_FILL
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER

    auto_col_widths(ws, min_width=15, max_width=60)
    ws.freeze_panes = "A3"
    return ws


def build_metadata_sheet(wb, scenario_data):
    ws = wb.create_sheet("Metadata")

    headers = ["Field", "Value"]
    ws.append(headers)
    apply_header_style(ws, 2)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    py_version = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"

    data_sources = (
        "results/scenario_comparison.csv, "
        "results/credit_ratings.csv, "
        "results/cashflow_*.csv (11 files), "
        "data/raw/plant_parameters.csv, "
        "src/climada/literature_parameters.py"
    )

    metadata = [
        ("Generation Timestamp", timestamp),
        ("Model Version", "CRP v2.0 (Feb 2026)"),
        ("Data Sources", data_sources),
        ("Number of Scenarios", len(scenario_data)),
        ("Python Version", py_version),
        ("Generator Script", "scripts/generate_excel_factcheck.py"),
        ("Output File", "results/CRP_Fact_Check_Workbook.xlsx"),
    ]

    for field, value in metadata:
        ws.append([field, value])

    nrows = ws.max_row
    for r in range(2, nrows + 1):
        fill = ALT_FILL if (r - 2) % 2 == 1 else WHITE_FILL
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER
            if c == 1:
                cell.font = BOLD_FONT

    auto_col_widths(ws, min_width=20, max_width=80)
    freeze_header(ws)
    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Loading data...")

    # Read main CSVs
    scenario_data = read_csv(SCENARIO_CSV)
    credit_data = read_csv(CREDIT_CSV)
    plant_data = read_csv(PLANT_CSV)

    # Read all cashflow CSVs
    cashflow_data = {}
    for sc, path in CASHFLOW_FILES.items():
        if os.path.exists(path):
            cashflow_data[sc] = read_csv(path)
        else:
            print(f"  Warning: {path} not found, skipping {sc}")
            cashflow_data[sc] = []

    print(f"  Loaded {len(scenario_data)} scenarios")
    print(f"  Loaded {len(credit_data)} credit rating rows")
    print(f"  Loaded {len(cashflow_data)} cashflow files")

    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building sheets...")

    # Sheet 1: Summary
    build_summary_sheet(wb, scenario_data)
    print("  [1/7] Summary")

    # Sheet 2: Credit Ratings
    build_credit_ratings_sheet(wb, credit_data)
    print("  [2/7] Credit Ratings")

    # Sheet 3: CRP Analysis
    build_crp_analysis_sheet(wb, scenario_data)
    print("  [3/7] CRP Analysis")

    # Sheet 4: Cashflows - Key Scenarios
    build_cashflows_key_sheet(wb, cashflow_data)
    print("  [4/7] Cashflows - Key Scenarios")

    # Sheet 5: All Cashflows Index
    build_all_cashflows_index(wb, cashflow_data)
    print("  [5/7] All Cashflows Index")

    # Sheet 6: Key Assumptions
    build_assumptions_sheet(wb, plant_data)
    print("  [6/7] Key Assumptions")

    # Sheet 7: Metadata
    build_metadata_sheet(wb, scenario_data)
    print("  [7/7] Metadata")

    # Save
    wb.save(OUTPUT_PATH)
    file_size = os.path.getsize(OUTPUT_PATH)
    print(f"\nWorkbook saved to: {OUTPUT_PATH}")
    print(f"File size: {file_size:,} bytes ({file_size / 1024:.1f} KB)")
    print(f"Sheets: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
